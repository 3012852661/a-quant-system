from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
OUT = Path(__file__).resolve().parent
OUT.mkdir(parents=True, exist_ok=True)
INPUT = ROOT / "outputs/batch_screen_20260615/批量初筛_清洗结果_20260615.csv"


def f(value: object, default: float = 0.0) -> float:
    try:
        if value in (None, "", "-"):
            return default
        return float(value)
    except Exception:
        return default


def band_score(value: float, ideal_low: float, ideal_high: float, hard_low: float, hard_high: float) -> float:
    if ideal_low <= value <= ideal_high:
        return 100.0
    if value < ideal_low:
        if value <= hard_low:
            return 0.0
        return (value - hard_low) / (ideal_low - hard_low) * 100
    if value >= hard_high:
        return 0.0
    return (hard_high - value) / (hard_high - ideal_high) * 100


def clamp(value: float, low: float = 0.0, high: float = 100.0) -> float:
    return max(low, min(high, value))


def read_rows() -> list[dict[str, str]]:
    with INPUT.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def factor_scores(row: dict[str, str]) -> dict[str, object]:
    pct = f(row["涨跌幅%"])
    amount_yi = f(row["成交额亿元"])
    turnover = f(row["换手率%"])
    vr = f(row["量比"], 1.0)
    pe = f(row["PE"])
    pb = f(row["PB"])
    mcap = f(row["总市值亿元"])
    main_net = f(row["主力净流入万元"])
    initial = f(row["初筛分"])
    risk_tags = row.get("风险标签", "")
    data_missing = amount_yi <= 0 or not row.get("行业")

    momentum = band_score(pct, 1.0, 7.0, -4.0, 10.5)
    if pct >= 9.5:
        momentum = min(momentum, 82.0)
    liquidity_amount = band_score(amount_yi, 3.0, 60.0, 0.0, 180.0) if amount_yi > 0 else 42.0
    liquidity_turnover = band_score(turnover, 1.0, 8.0, 0.0, 18.0)
    liquidity = liquidity_amount * 0.62 + liquidity_turnover * 0.38

    if pe <= 0:
        valuation_pe = 25.0
    else:
        valuation_pe = band_score(pe, 5.0, 35.0, 0.0, 120.0)
    valuation_pb = band_score(pb, 0.8, 6.0, 0.0, 16.0) if pb > 0 else 45.0
    valuation = valuation_pe * 0.62 + valuation_pb * 0.38

    fund = 50.0
    if main_net >= 5000:
        fund = 92.0
    elif main_net > 0:
        fund = 70.0
    elif main_net <= -5000:
        fund = 18.0
    elif main_net < 0:
        fund = 35.0

    size_quality = band_score(mcap, 50.0, 1500.0, 15.0, 3500.0) if mcap > 0 else 55.0

    penalty = 0.0
    reasons: list[str] = []
    if "ST/*ST" in risk_tags:
        penalty += 42
        reasons.append("ST/*ST")
    if "亏损PE" in risk_tags:
        penalty += 18
        reasons.append("亏损PE")
    if "高换手" in risk_tags:
        penalty += 10
        reasons.append("高换手")
    if "大幅波动" in risk_tags:
        penalty += 12
        reasons.append("大幅波动")
    if data_missing:
        penalty += 8
        reasons.append("行情字段不完整")
    if row["名称不一致"] == "是" and row["代码"] != "000725":
        penalty += 5
        reasons.append("名称需核验")
    if vr >= 4:
        penalty += min(12, (vr - 4) * 3)
        reasons.append("量比过热")

    total = (
        momentum * 0.24
        + liquidity * 0.20
        + valuation * 0.16
        + fund * 0.16
        + size_quality * 0.10
        + initial * 0.14
        - penalty
    )
    total = clamp(total)

    if "ST/*ST" in risk_tags:
        layer = "R-风险剔除"
    elif total >= 78:
        layer = "Q1-重点量化研究"
    elif total >= 68:
        layer = "Q2-观察研究"
    elif total >= 55:
        layer = "Q3-低优先级跟踪"
    else:
        layer = "R-风险剔除"
    if layer.startswith("R") and not reasons:
        reasons.append("综合分低/因子不占优")

    return {
        "量化总分": round(total, 2),
        "研究层级": layer,
        "动量分": round(momentum, 1),
        "流动性分": round(liquidity, 1),
        "估值分": round(valuation, 1),
        "资金分": round(fund, 1),
        "市值适配分": round(size_quality, 1),
        "风险扣分": round(penalty, 1),
        "量化风险原因": "、".join(reasons),
        "数据完整性": "不完整" if data_missing else "完整",
    }


def build_analysis() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for raw in read_rows():
        factors = factor_scores(raw)
        rows.append({**raw, **factors})
    layer_order = {"Q1-重点量化研究": 0, "Q2-观察研究": 1, "Q3-低优先级跟踪": 2, "R-风险剔除": 3}
    rows.sort(key=lambda r: (layer_order.get(str(r["研究层级"]), 9), -f(r["量化总分"])))
    return rows


def industry_summary(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    groups: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        industry = str(row.get("行业") or "未分类")
        groups[industry].append(row)
    result = []
    for industry, items in groups.items():
        q1 = [r for r in items if str(r["研究层级"]).startswith("Q1")]
        avg_score = sum(f(r["量化总分"]) for r in items) / len(items)
        avg_pct = sum(f(r["涨跌幅%"]) for r in items) / len(items)
        amount = sum(f(r["成交额亿元"]) for r in items)
        top = sorted(items, key=lambda r: f(r["量化总分"]), reverse=True)[0]
        result.append({
            "行业": industry,
            "数量": len(items),
            "Q1数量": len(q1),
            "平均量化分": round(avg_score, 2),
            "平均涨跌幅%": round(avg_pct, 2),
            "合计成交额亿元": round(amount, 2),
            "行业最高分标的": f"{top['代码']} {top['真实简称']}",
            "最高分": top["量化总分"],
        })
    result.sort(key=lambda r: (r["Q1数量"], r["平均量化分"], r["合计成交额亿元"]), reverse=True)
    return result


def write_csv(rows: list[dict[str, object]]) -> Path:
    path = OUT / "量化分析全表_20260615.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return path


def style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4D78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color="D9E2F3"))


def write_xlsx(rows: list[dict[str, object]], industries: list[dict[str, object]]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "量化全表"
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for name, predicate in [
        ("Q1重点研究", lambda r: str(r["研究层级"]).startswith("Q1")),
        ("Q2观察研究", lambda r: str(r["研究层级"]).startswith("Q2")),
        ("风险剔除", lambda r: str(r["研究层级"]).startswith("R")),
    ]:
        sheet = wb.create_sheet(name)
        sheet.append(headers)
        for row in rows:
            if predicate(row):
                sheet.append([row.get(h, "") for h in headers])

    ws_ind = wb.create_sheet("行业聚合")
    ind_headers = list(industries[0].keys()) if industries else ["行业"]
    ws_ind.append(ind_headers)
    for row in industries:
        ws_ind.append([row.get(h, "") for h in ind_headers])

    ws_note = wb.create_sheet("模型说明")
    notes = [
        ["字段", "说明"],
        ["量化总分", "动量24% + 流动性20% + 估值16% + 资金16% + 市值适配10% + 初筛分14% - 风险扣分"],
        ["研究层级", "Q1/Q2/Q3/R 代表研究优先级，不是交易指令"],
        ["数据完整性", "成交额或行业缺失时标记不完整，并额外扣分"],
        ["风险扣分", "ST/*ST、亏损PE、高换手、大幅波动、名称需核验、行情字段不完整、量比过热"],
        ["数据日期", "2026-06-15，来自批量清洗阶段的东方财富/腾讯公开行情快照"],
    ]
    for item in notes:
        ws_note.append(item)

    for sheet in wb.worksheets:
        style_sheet(sheet)
        for idx in range(1, min(sheet.max_column, 32) + 1):
            width = 12
            if idx in (1, 4, 16, 17, 18, 21, 31):
                width = 22
            if idx in (17, 18):
                width = 40
            sheet.column_dimensions[get_column_letter(idx)].width = width

    path = OUT / "量化分析结果_20260615.xlsx"
    wb.save(path)
    return path


def write_md(rows: list[dict[str, object]], industries: list[dict[str, object]]) -> Path:
    q1 = [r for r in rows if str(r["研究层级"]).startswith("Q1")]
    q2 = [r for r in rows if str(r["研究层级"]).startswith("Q2")]
    risk = [r for r in rows if str(r["研究层级"]).startswith("R")]
    path = OUT / "量化分析摘要_20260615.md"
    lines = [
        "# 量化分析摘要",
        "",
        "声明：本文件仅用于研究和产品能力展示，不构成任何真实投资建议；研究层级不是买卖信号。",
        "",
        f"- 分析股票数：{len(rows)}",
        f"- Q1重点量化研究：{len(q1)}",
        f"- Q2观察研究：{len(q2)}",
        f"- 风险剔除：{len(risk)}",
        "",
        "## Q1 Top 20",
        "",
    ]
    for row in q1[:20]:
        lines.append(
            f"- {row['代码']} {row['真实简称']}：{row['量化总分']}分，行业={row.get('行业') or '未分类'}，"
            f"涨跌幅={row['涨跌幅%']}%，成交额={row['成交额亿元']}亿元，"
            f"动量/流动性/估值/资金={row['动量分']}/{row['流动性分']}/{row['估值分']}/{row['资金分']}，"
            f"风险={row.get('量化风险原因') or '无'}"
        )
    lines += ["", "## 行业聚合 Top 10", ""]
    for row in industries[:10]:
        lines.append(
            f"- {row['行业']}：数量{row['数量']}，Q1数量{row['Q1数量']}，平均分{row['平均量化分']}，"
            f"成交额合计{row['合计成交额亿元']}亿元，最高分={row['行业最高分标的']}"
        )
    lines += ["", "## 风险剔除样本", ""]
    for row in risk[:25]:
        lines.append(f"- {row['代码']} {row['真实简称']}：{row['量化总分']}分，原因={row.get('量化风险原因') or row.get('风险标签')}")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


if __name__ == "__main__":
    rows = build_analysis()
    industries = industry_summary(rows)
    print(write_csv(rows))
    print(write_xlsx(rows, industries))
    print(write_md(rows, industries))
